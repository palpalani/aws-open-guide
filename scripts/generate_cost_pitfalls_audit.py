#!/usr/bin/env python3
"""Regenerate use-cases/cost-pitfalls-audit.xlsx from use-cases/cost-pitfalls.md.

Run: uv run --python python3 --with openpyxl python3 scripts/generate_cost_pitfalls_audit.py
Optional: LibreOffice (for formula recalculation via the document-skills recalc helper)
"""

from __future__ import annotations

import re
import subprocess
import sys
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_MD = REPO_ROOT / "use-cases" / "cost-pitfalls.md"
OUTPUT_XLSX = REPO_ROOT / "use-cases" / "cost-pitfalls-audit.xlsx"

SKIP_SECTIONS = frozenset(
    {
        "Frequently asked questions",
        "Quarterly optimization cadence",
        "Tools to find these proactively",
        "Cost discipline checklist",
    }
)

QUARTERLY_STEP: dict[str, str] = {
    "NAT Gateway": "Architecture",
    "Cross-AZ data transfer": "Architecture",
    "CloudWatch Logs": "Governance",
    "Egress to internet": "Architecture",
    "DynamoDB hot partitions and scan-heavy tables": "Rightsizing",
    "S3 small-object PUT-heavy without aggregation": "Architecture",
    "EBS gp2 vs gp3 (almost free win)": "Rightsizing",
    "Idle resources": "Waste",
    "Lambda over-provisioned memory": "Rightsizing",
    "Reserved capacity and Savings Plans": "Commitments",
    "Cross-region traffic": "Architecture",
    "Free-tier-as-DoS": "Governance",
    "Bedrock and GenAI tokens": "Rightsizing",
}

URL_RE = re.compile(r"https?://[^\s)]+")


def _collapse(text: str, limit: int = 220) -> str:
    one_line = " ".join(line.strip() for line in text.strip().splitlines() if line.strip())
    one_line = re.sub(r"\s+", " ", one_line)
    if len(one_line) <= limit:
        return one_line
    return one_line[: limit - 1].rstrip() + "…"


def _extract_block(body: str, *labels: str) -> str:
    for label in labels:
        pattern = rf"\*\*{re.escape(label)}:\*\*\s*\n(.*?)(?=\n\*\*|\n---|\Z)"
        match = re.search(pattern, body, re.DOTALL)
        if match:
            return _collapse(match.group(1))
    return ""


def _first_url(body: str) -> str:
    ref = re.search(r"\*\*Reference:\*\*\s*(.+)", body)
    if ref:
        urls = URL_RE.findall(ref.group(1))
        if urls:
            return urls[0].rstrip(".")
    return ""


def _bullets_summary(body: str, label: str, max_items: int = 2) -> str:
    block = _extract_block(body, label)
    if not block:
        return ""
    items = re.findall(r"^- (.+)$", block, re.MULTILINE)
    if not items:
        return block
    cleaned = []
    for item in items[:max_items]:
        cleaned.append(re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", item).strip())
    return "; ".join(cleaned)


def parse_markdown(text: str) -> tuple[list[tuple], list[str], list[tuple]]:
    sections: dict[str, str] = {}
    for chunk in re.split(r"\n## ", text)[1:]:
        title, _, body = chunk.partition("\n")
        sections[title.strip()] = body

    pitfalls: list[tuple] = []
    for title, body in sections.items():
        if title in SKIP_SECTIONS:
            continue
        pitfalls.append(
            (
                title,
                _bullets_summary(body, "The cost", 2) or _extract_block(body, "The cost"),
                _extract_block(body, "Why it spirals", "Why teams miss it"),
                _bullets_summary(body, "Mitigation", 3),
                _first_url(body),
                QUARTERLY_STEP.get(title, ""),
            )
        )

    checklist_body = sections.get("Cost discipline checklist", "")
    checklist = re.findall(r"- \[ \] (.+)", checklist_body)

    cadence_body = sections.get("Quarterly optimization cadence", "")
    cadence: list[tuple] = []
    for match in re.finditer(r"^(\d+)\. \*\*([^*]+)\*\* — (.+)$", cadence_body, re.MULTILINE):
        step, focus, actions = match.groups()
        section = re.search(r"\(#([^)]+)\)", actions)
        cadence.append((step, focus.strip(), _collapse(actions, 280), section.group(1) if section else focus.strip()))

    return pitfalls, checklist, cadence


def build_workbook(pitfalls: list[tuple], checklist: list[str], cadence: list[tuple]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required: uv run --python python3 --with openpyxl python3 scripts/generate_cost_pitfalls_audit.py"
        ) from exc

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color="232F3E")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers: list[str]) -> None:
        for col, label in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    ws = wb.active
    ws.title = "Pitfalls Registry"
    headers = ["Pitfall", "Primary Cost Driver", "Why It Spirals", "Top Mitigation", "AWS Reference", "Quarterly Step"]
    style_header(ws, headers)
    for row_idx, row in enumerate(pitfalls, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = wrap
            cell.border = border
            if col_idx == 5 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
    for idx, width in enumerate([22, 28, 38, 38, 42, 14], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    if pitfalls:
        ws.auto_filter.ref = f"A1:F{1 + len(pitfalls)}"

    ws2 = wb.create_sheet("Audit Checklist")
    check_headers = ["#", "Checklist Item", "Owner", "Status", "Notes", "Source Section"]
    style_header(ws2, check_headers)
    dv = DataValidation(type="list", formula1='"Not started,In progress,Done,N/A"', allow_blank=True)
    ws2.add_data_validation(dv)
    last_row = 1
    for i, item in enumerate(checklist, 1):
        last_row = i + 1
        ws2.cell(row=last_row, column=1, value=i).border = border
        item_cell = ws2.cell(row=last_row, column=2, value=item)
        item_cell.border = border
        item_cell.alignment = wrap
        item_cell.font = Font(name="Arial", size=10)
        ws2.cell(row=last_row, column=3, value="").border = border
        status_cell = ws2.cell(row=last_row, column=4, value="Not started")
        status_cell.border = border
        status_cell.font = Font(name="Arial", size=10)
        dv.add(status_cell)
        ws2.cell(row=last_row, column=5, value="").border = border
        src = ws2.cell(row=last_row, column=6, value="cost-pitfalls.md")
        src.border = border
        src.font = Font(name="Arial", size=10, color="666666")

    summary_row = last_row + 2
    ws2.cell(row=summary_row, column=1, value="Progress").font = Font(bold=True, name="Arial")
    ws2.cell(row=summary_row, column=2, value=f'=COUNTIF(D2:D{last_row},"Done")&" of "&COUNTA(B2:B{last_row})&" complete"')
    ws2.cell(row=summary_row, column=2).font = Font(bold=True, name="Arial", size=11)
    ws2.cell(row=summary_row + 1, column=1, value="% Complete").font = Font(bold=True, name="Arial")
    ws2.cell(
        row=summary_row + 1,
        column=2,
        value=f'=IF(COUNTA(B2:B{last_row})=0,"-",ROUND(COUNTIF(D2:D{last_row},"Done")/COUNTA(B2:B{last_row}),2))',
    )
    ws2.cell(row=summary_row + 1, column=2).number_format = "0%"
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 62
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 28
    ws2.column_dimensions["F"].width = 18
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Quarterly Cadence")
    cadence_headers = ["Step", "Focus Area", "Key Actions", "Playbook Section"]
    style_header(ws3, cadence_headers)
    for row_idx, row in enumerate(cadence, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = wrap
            cell.border = border
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 32
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet("About")
    about = [
        ("Source", "use-cases/cost-pitfalls.md"),
        ("Generated by", "scripts/generate_cost_pitfalls_audit.py"),
        ("Purpose", "Quarterly AWS cost audit workbook — filter pitfalls, track checklist, follow cadence"),
        ("Pricing note", "Verify all dollar figures on linked AWS pricing pages before committing"),
        ("Last synced", date.today().isoformat()),
    ]
    for i, (key, value) in enumerate(about, 1):
        ws4.cell(row=i, column=1, value=key).font = Font(bold=True, name="Arial")
        ws4.cell(row=i, column=2, value=value).font = Font(name="Arial")
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 70

    wb.save(OUTPUT_XLSX)


def maybe_recalc() -> None:
    recalc = Path.home() / ".cursor/plugins/cache/anthropic-agent-skills/document-skills/1ed29a03dc852d30fa6ef2ca53a67dc2c2c2c563/skills/xlsx/scripts/recalc.py"
    if not recalc.is_file():
        return
    subprocess.run([sys.executable, str(recalc), str(OUTPUT_XLSX), "30"], check=False)


def main() -> int:
    if not SOURCE_MD.is_file():
        print(f"Missing source: {SOURCE_MD}", file=sys.stderr)
        return 1
    text = SOURCE_MD.read_text(encoding="utf-8")
    pitfalls, checklist, cadence = parse_markdown(text)
    if not pitfalls:
        print("No pitfall sections parsed — check cost-pitfalls.md structure.", file=sys.stderr)
        return 1
    build_workbook(pitfalls, checklist, cadence)
    maybe_recalc()
    print(f"Wrote {OUTPUT_XLSX.relative_to(REPO_ROOT)} ({len(pitfalls)} pitfalls, {len(checklist)} checklist items)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
